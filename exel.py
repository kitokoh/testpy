import sys
import os
from PyQt5.QtWidgets import (QApplication, QMainWindow, QTableView, QFileDialog,
                             QVBoxLayout, QWidget, QToolBar, QAction, QStatusBar,
                             QMessageBox, QLabel)
from PyQt5.QtCore import Qt, QMimeData
from PyQt5.QtGui import QStandardItemModel, QStandardItem, QIcon, QFont, QColor
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelEditor(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Éditeur Excel Professionnel")
        self.setGeometry(100, 100, 1200, 800)
        
        # Variables
        self.current_file = None
        self.data_model = None
        self.workbook = None
        self.sheet = None
        
        # Configuration du style
        self.setup_styles()
        
        # Création de l'interface
        self.create_ui()
        
        # Activer le glisser-déposer
        self.setAcceptDrops(True)
        
    def setup_styles(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QToolBar {
                background-color: #e1e1e1;
                border-bottom: 1px solid #ccc;
                padding: 2px;
            }
            QToolButton {
                padding: 5px;
            }
            QTableView {
                background-color: white;
                alternate-background-color: #f9f9f9;
                gridline-color: #ddd;
                font-size: 12px;
            }
            QTableView::item:selected {
                background-color: #4a90e2;
                color: white;
            }
            QStatusBar {
                background-color: #e1e1e1;
                color: #555;
                font-size: 11px;
            }
            QLabel#dropLabel {
                font-size: 18px;
                color: #666;
                qproperty-alignment: AlignCenter;
                border: 2px dashed #aaa;
                padding: 50px;
                margin: 20px;
            }
        """)
    
    def create_ui(self):
        # Widget central et layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        self.layout = QVBoxLayout(central_widget)
        self.layout.setContentsMargins(10, 10, 10, 10)
        
        # Label pour le glisser-déposer
        self.drop_label = QLabel("Glissez-déposez un fichier Excel ici\nou utilisez le menu Fichier")
        self.drop_label.setObjectName("dropLabel")
        self.drop_label.setFont(QFont("Arial", 14, QFont.Bold))
        self.layout.addWidget(self.drop_label)
        
        # Table view pour afficher les données Excel
        self.table_view = QTableView()
        self.table_view.setSortingEnabled(True)
        self.table_view.hide()  # Caché au démarrage
        self.layout.addWidget(self.table_view)
        
        # Création de la barre d'outils
        self.create_toolbar()
        
        # Barre de statut
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Prêt")
    
    def create_toolbar(self):
        toolbar = QToolBar("Barre d'outils principale")
        self.addToolBar(toolbar)
        
        # Actions avec icônes
        open_action = QAction(QIcon.fromTheme("document-open"), "Ouvrir", self)
        open_action.triggered.connect(self.open_file)
        open_action.setShortcut("Ctrl+O")
        toolbar.addAction(open_action)
        
        save_action = QAction(QIcon.fromTheme("document-save"), "Enregistrer", self)
        save_action.triggered.connect(self.save_file)
        save_action.setShortcut("Ctrl+S")
        toolbar.addAction(save_action)
        
        save_as_action = QAction(QIcon.fromTheme("document-save-as"), "Enregistrer sous...", self)
        save_as_action.triggered.connect(self.save_file_as)
        toolbar.addAction(save_as_action)
        
        toolbar.addSeparator()
        
        quit_action = QAction(QIcon.fromTheme("application-exit"), "Quitter", self)
        quit_action.triggered.connect(self.close)
        quit_action.setShortcut("Ctrl+Q")
        toolbar.addAction(quit_action)
    
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
    
    def dropEvent(self, event):
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.lower().endswith(('.xlsx', '.xlsm')):
                self.load_excel_file(file_path)
                break
        else:
            QMessageBox.warning(self, "Format non supporté", 
                              "Seuls les fichiers Excel (.xlsx, .xlsm) sont supportés.")
    
    def load_excel_file(self, file_path):
        try:
            # Charger le workbook avec openpyxl
            self.workbook = load_workbook(filename=file_path)
            self.sheet = self.workbook.active
            
            # Créer le modèle de données
            self.data_model = QStandardItemModel(self.sheet.max_row, self.sheet.max_column)
            
            # Remplir le modèle avec les données et les styles
            for row in range(1, self.sheet.max_row + 1):
                for col in range(1, self.sheet.max_column + 1):
                    cell = self.sheet.cell(row=row, column=col)
                    item = QStandardItem()
                    
                    # Valeur de la cellule
                    if cell.value is not None:
                        if isinstance(cell.value, str):
                            item.setText(cell.value)
                        else:
                            # Pour les nombres, dates, etc.
                            item.setData(cell.value, Qt.DisplayRole)
                    
                    # Style de la cellule
                    self.apply_cell_style(item, cell)
                    
                    self.data_model.setItem(row - 1, col - 1, item)
            
            # Configurer les en-têtes de colonnes
            for col in range(1, self.sheet.max_column + 1):
                letter = get_column_letter(col)
                self.data_model.setHeaderData(col - 1, Qt.Horizontal, letter)
            
            # Appliquer le modèle à la table view
            self.table_view.setModel(self.data_model)
            
            # Ajuster la largeur des colonnes
            for col in range(self.sheet.max_column):
                self.table_view.setColumnWidth(col, 120)  # Largeur par défaut
                
                # Essayer d'utiliser la largeur de colonne d'Excel si disponible
                try:
                    col_letter = get_column_letter(col + 1)
                    if self.sheet.column_dimensions[col_letter].width:
                        width = int(self.sheet.column_dimensions[col_letter].width * 7)  # Facteur d'échelle
                        self.table_view.setColumnWidth(col, width)
                except:
                    pass
            
            self.table_view.show()
            self.drop_label.hide()
            
            # Mettre à jour l'état
            self.current_file = file_path
            self.status_bar.showMessage(f"Fichier chargé: {os.path.basename(file_path)}")
            self.setWindowTitle(f"Éditeur Excel - {os.path.basename(file_path)}")
            
        except Exception as e:
            QMessageBox.critical(
                self, "Erreur", 
                f"Impossible d'ouvrir le fichier:\n{str(e)}"
            )
    
    def apply_cell_style(self, item, cell):
        # Police
        if cell.font:
            font = QFont()
            if cell.font.bold:
                font.setBold(True)
            if cell.font.italic:
                font.setItalic(True)
            if cell.font.underline:
                font.setUnderline(True)
            if cell.font.strike:
                font.setStrikeOut(True)
            if cell.font.size:
                try:
                    # Convertir la taille de police en entier
                    font_size = int(round(cell.font.size))
                    font.setPointSize(font_size)
                except:
                    pass
            if cell.font.name:
                font.setFamily(cell.font.name)
            item.setFont(font)
        
        # Couleur de fond
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
            try:
                color = QColor(cell.fill.start_color.rgb)
                item.setBackground(color)
            except:
                pass
        
        # Couleur du texte
        if cell.font and cell.font.color and cell.font.color.rgb:
            try:
                color = QColor(cell.font.color.rgb)
                item.setForeground(color)
            except:
                pass
        
        # Alignement
        if cell.alignment:
            alignment = Qt.Alignment()
            if cell.alignment.horizontal == 'left':
                alignment |= Qt.AlignLeft
            elif cell.alignment.horizontal == 'center':
                alignment |= Qt.AlignHCenter
            elif cell.alignment.horizontal == 'right':
                alignment |= Qt.AlignRight
            
            if cell.alignment.vertical == 'top':
                alignment |= Qt.AlignTop
            elif cell.alignment.vertical == 'center':
                alignment |= Qt.AlignVCenter
            elif cell.alignment.vertical == 'bottom':
                alignment |= Qt.AlignBottom
            
            item.setTextAlignment(alignment)
    
    def open_file(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Ouvrir un fichier Excel", "", 
            "Fichiers Excel (*.xlsx *.xlsm);;Tous les fichiers (*)", 
            options=options
        )
        
        if file_path:
            self.load_excel_file(file_path)
    
    def save_file(self):
        if not self.current_file or not self.workbook:
            self.save_file_as()
            return
            
        try:
            # Mettre à jour les données dans le workbook
            for row in range(self.data_model.rowCount()):
                for col in range(self.data_model.columnCount()):
                    item = self.data_model.item(row, col)
                    if item:
                        cell = self.sheet.cell(row=row + 1, column=col + 1)
                        cell.value = item.data(Qt.DisplayRole)
            
            # Sauvegarder le fichier
            self.workbook.save(self.current_file)
            self.status_bar.showMessage(f"Fichier enregistré: {os.path.basename(self.current_file)}")
            
        except Exception as e:
            QMessageBox.critical(
                self, "Erreur", 
                f"Impossible d'enregistrer le fichier:\n{str(e)}"
            )
    
    def save_file_as(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer le fichier", "", 
            "Fichiers Excel (*.xlsx);;Tous les fichiers (*)", 
            options=options
        )
        
        if file_path:
            # Ajouter l'extension si elle manque
            if not file_path.lower().endswith('.xlsx'):
                file_path += '.xlsx'
            
            self.current_file = file_path
            self.save_file()
    
    def closeEvent(self, event):
        reply = QMessageBox.question(
            self, 'Quitter',
            "Voulez-vous vraiment quitter?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # Configuration pour améliorer l'apparence
    app.setStyle('Fusion')
    
    # Définir une police par défaut
    font = QFont()
    font.setFamily("Segoe UI" if sys.platform == "win32" else "Arial")
    font.setPointSize(10)
    app.setFont(font)
    
    editor = ExcelEditor()
    editor.show()
    
    sys.exit(app.exec_())