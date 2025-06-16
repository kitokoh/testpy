"""
Ultimate Excel Editor - Solution complète pour l'édition et l'export PDF de modèles Excel complexes
"""

import os
import sys
import logging
import tempfile
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Union
from enum import Enum
from dataclasses import dataclass
from pathlib import Path

# Excel handling
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Fill, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException, ReadOnlyWorkbookException
import zipfile
from openpyxl.worksheet.page import PageMargins

# PDF export
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.lib.units import cm, inch, mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus.flowables import Flowable
import io # Added for BytesIO
from reportlab.lib.utils import ImageReader # Added for header/footer images

# Qt GUI
from PyQt5.QtWidgets import (
    QApplication, QDialog, QTableWidget, QTableWidgetItem, QMessageBox,
    QVBoxLayout, QHBoxLayout, QPushButton, QDialogButtonBox,
    QAbstractItemView, QHeaderView, QFileDialog, QProgressBar,
    QLabel, QFrame, QSplitter, QWidget, QGroupBox, QFormLayout,
    QLineEdit, QTextEdit, QComboBox, QSpinBox, QDoubleSpinBox, QCheckBox,
    QStyleFactory, QStyle, QSizePolicy, QScrollArea, QToolBar, QAction
)
from PyQt5.QtGui import (
    QFont, QColor, QIcon, QPalette, QBrush, QKeySequence,
    QTextDocument, QTextCursor, QTextCharFormat, QTextTableFormat,
    QTextLength, QSyntaxHighlighter, QTextFormat
)
from PyQt5.QtCore import (
    Qt, QThread, pyqtSignal, QTimer, QCoreApplication, QSize,
    QRectF, QEvent, QObject, QFileInfo, QSettings
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Constants
class PageOrientation(Enum):
    PORTRAIT = 1
    LANDSCAPE = 2

DEFAULT_FONT_NAME = "Arial" # This is for Excel cell style, not PDF
MAX_ROWS_PREVIEW = 5000
MAX_COLS_PREVIEW = 100
PDF_DPI = 300
PDF_PAGE_MARGINS = 15 * mm
PDF_HEADER_HEIGHT = 15 * mm
PDF_FOOTER_HEIGHT = 10 * mm

@dataclass
class ExcelCellStyle:
    font_name: str = "Calibri"
    font_size: int = 11
    bold: bool = False
    italic: bool = False
    underline: bool = False
    text_color: str = "#000000"
    bg_color: str = "#FFFFFF"
    h_align: str = "left"
    v_align: str = "center"
    border: str = "none"
    wrap_text: bool = False
    number_format: str = "General"

@dataclass
class ClientData:
    name: str = ""
    company: str = ""
    address: str = ""
    phone: str = ""
    email: str = ""
    project: str = ""
    project_id: str = ""
    price: float = 0.0
    currency: str = "€"
    notes: str = ""
    logo_path: str = ""

@dataclass
class PDFExportSettings:
    orientation: PageOrientation = PageOrientation.PORTRAIT
    page_size: Tuple[float, float] = A4
    margins: Tuple[float, float, float, float] = (PDF_PAGE_MARGINS, PDF_PAGE_MARGINS, PDF_PAGE_MARGINS, PDF_PAGE_MARGINS)
    header: bool = True
    footer: bool = True
    grid_lines: bool = True
    repeat_headers: bool = True
    watermark: str = ""
    quality: int = 100  # %

class ExcelTableModel:
    """Represents the complete Excel workbook model"""
    
    def __init__(self):
        self.file_path: Optional[str] = None
        self.workbook: Optional[Workbook] = None
        self.current_sheet: Optional[Worksheet] = None
        self.sheets: List[str] = []
        self.client_data: ClientData = ClientData()
        self.is_modified: bool = False
        self.load_error_message: Optional[str] = None
        self.sheet_images: Dict[str, List] = {}  # To store images per sheet
        self.sheet_headers_footers: Dict[str, Dict[str, Dict[str, Dict[str, Optional[str]]]]] = {} # To store headers/footers
    
    def load_workbook(self, file_path: str):
        """Load workbook from file with maximum compatibility"""
        self.load_error_message = None
        self.workbook = None
        try:
            logger.info(f"Attempting to load workbook (styles preserved): {file_path}")
            self.workbook = load_workbook(
                file_path,
                read_only=False,
                keep_vba=True,
                data_only=False,
                keep_links=True
            )
            logger.info(f"Successfully loaded {file_path} with styles (read-write).")

        except (InvalidFileException, ReadOnlyWorkbookException, IOError, zipfile.BadZipFile) as e_rw:
            logger.warning(f"Read-write load attempt failed for {file_path} with {type(e_rw).__name__}: {e_rw}. Trying read-only.")
            try:
                logger.info(f"Attempting to load workbook (styles preserved, read-only): {file_path}")
                self.workbook = load_workbook(
                    file_path,
                    read_only=True,
                    keep_vba=True,
                    data_only=False,
                    keep_links=True
                )
                logger.info(f"Successfully loaded {file_path} with styles (read-only).")
            except (InvalidFileException, ReadOnlyWorkbookException, IOError, zipfile.BadZipFile) as e_ro:
                error_msg = f"Error loading workbook '{file_path}' (read-only attempt): {type(e_ro).__name__} - {str(e_ro)}"
                logger.error(error_msg)
                self.load_error_message = error_msg
                self.workbook = None
                return False
            except Exception as e_generic_ro:
                error_msg = f"Unexpected error loading workbook '{file_path}' (read-only attempt): {type(e_generic_ro).__name__} - {str(e_generic_ro)}"
                logger.error(error_msg)
                self.load_error_message = error_msg
                self.workbook = None
                return False
        except Exception as e_generic_rw:
            error_msg = f"Unexpected error loading workbook '{file_path}': {type(e_generic_rw).__name__} - {str(e_generic_rw)}"
            logger.error(error_msg)
            self.load_error_message = error_msg
            self.workbook = None
            return False

        if self.workbook:
            self.file_path = file_path
            self.sheets = self.workbook.sheetnames
            self.current_sheet = self.workbook.active
            self.is_modified = False
            self.sheet_images.clear()
            self.sheet_headers_footers.clear()

            for sheet in self.workbook.worksheets:
                if sheet._images: # Corrected attribute
                    self.sheet_images[sheet.title] = list(sheet._images)
                    logger.info(f"Found {len(self.sheet_images[sheet.title])} image(s) in sheet '{sheet.title}'.")
                else:
                    logger.info(f"No images found in sheet '{sheet.title}'.")

                hf_data = {}
                if sheet.HeaderFooter: # Corrected attribute
                    hf_obj_main = sheet.HeaderFooter
                    hf_types = {
                        "odd_header": hf_obj_main.oddHeader,
                        "even_header": hf_obj_main.evenHeader,
                        "first_header": hf_obj_main.firstHeader,
                        "odd_footer": hf_obj_main.oddFooter,
                        "even_footer": hf_obj_main.evenFooter,
                        "first_footer": hf_obj_main.firstFooter,
                    }
                    for hf_type, hf_member_obj in hf_types.items():
                        if hf_member_obj:
                            left_text = hf_member_obj.left.text if hf_member_obj.left else None
                            center_text = hf_member_obj.center.text if hf_member_obj.center else None
                            right_text = hf_member_obj.right.text if hf_member_obj.right else None

                            hf_data[hf_type] = {
                                "left": {"text": left_text},
                                "center": {"text": center_text},
                                "right": {"text": right_text},
                            }
                            log_msg = f"Extracted {hf_type} for sheet '{sheet.title}':"
                            if left_text: log_msg += f" L='{left_text}'"
                            if center_text: log_msg += f" C='{center_text}'"
                            if right_text: log_msg += f" R='{right_text}'"
                            logger.info(log_msg)

                            if left_text and "&G" in left_text:
                                logger.info(f"  Found image reference mark (&G) in {hf_type} left for sheet '{sheet.title}'.")
                            if center_text and "&G" in center_text:
                                logger.info(f"  Found image reference mark (&G) in {hf_type} center for sheet '{sheet.title}'.")
                            if right_text and "&G" in right_text:
                                logger.info(f"  Found image reference mark (&G) in {hf_type} right for sheet '{sheet.title}'.")

                if hf_data:
                    self.sheet_headers_footers[sheet.title] = hf_data
                else:
                    logger.info(f"No header/footer data found for sheet '{sheet.title}'.")

            self.extract_client_data()
            return True

        if not self.workbook:
            if not self.load_error_message:
                 self.load_error_message = f"Failed to load workbook '{file_path}' after multiple attempts."
            logger.error(self.load_error_message)
            return False
        return False
    
    def extract_client_data(self):
        if not self.current_sheet:
            return
        patterns = {
            "name": ["client", "customer", "nom", "name"],
            "company": ["company", "société", "entreprise"],
            "project": ["project", "projet", "description"],
            "price": ["price", "prix", "cost", "montant"]
        }
        for row_idx, row in enumerate(self.current_sheet.iter_rows(max_row=20)): # Added row_idx for clarity
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    lower_val = cell.value.lower()
                    if any(p in lower_val for p in patterns["name"]):
                        try:
                            self.client_data.name = str(self.current_sheet.cell(row=cell.row, column=cell.column+1).value)
                        except: pass
                    elif any(p in lower_val for p in patterns["company"]):
                        try:
                            self.client_data.company = str(self.current_sheet.cell(row=cell.row, column=cell.column+1).value)
                        except: pass
                    elif any(p in lower_val for p in patterns["project"]):
                        try:
                            self.client_data.project = str(self.current_sheet.cell(row=cell.row, column=cell.column+1).value)
                        except: pass
                    elif any(p in lower_val for p in patterns["price"]):
                        try:
                            val = self.current_sheet.cell(row=cell.row, column=cell.column+1).value
                            if isinstance(val, (int, float)):
                                self.client_data.price = float(val)
                        except: pass

# --- Start of StyleConverter ---
class StyleConverter:
    """Handles conversion between Excel, Qt and PDF styles"""
    
    @staticmethod
    def excel_to_qt(excel_cell) -> ExcelCellStyle:
        """Convert Excel cell style to Qt compatible style"""
        style = ExcelCellStyle()
        
        if excel_cell.has_style:
            # Font properties
            if excel_cell.font:
                style.font_name = excel_cell.font.name or "Calibri"
                style.font_size = excel_cell.font.sz or 11
                style.bold = excel_cell.font.b
                style.italic = excel_cell.font.i
                style.underline = excel_cell.font.u
                if excel_cell.font.color and excel_cell.font.color.rgb:
                    style.text_color = f"#{str(excel_cell.font.color.rgb)[2:]}"
            
            # Background color
            if excel_cell.fill and excel_cell.fill.fgColor and excel_cell.fill.fgColor.rgb:
                style.bg_color = f"#{str(excel_cell.fill.fgColor.rgb)[2:]}"
            
            # Alignment
            if excel_cell.alignment:
                style.h_align = excel_cell.alignment.horizontal or "left"
                style.v_align = excel_cell.alignment.vertical or "center"
                style.wrap_text = excel_cell.alignment.wrapText
            
            # Number format
            style.number_format = excel_cell.number_format or "General"
        
        return style
    
    @staticmethod
    def qt_to_excel(qt_item, excel_cell):
        """Apply Qt item style to Excel cell"""
        if not qt_item or not excel_cell:
            return
            
        # Font
        font = Font(
            name=str(qt_item.font().family()),
            sz=qt_item.font().pointSize(),
            b=qt_item.font().bold(),
            i=qt_item.font().italic(),
            u=qt_item.font().underline(),
            color=StyleConverter.hex_to_excel_rgb(qt_item.foreground().color().name())
        )
        excel_cell.font = font
        
        # Background
        fill = PatternFill(
            start_color=StyleConverter.hex_to_excel_rgb(qt_item.background().color().name()),
            end_color=StyleConverter.hex_to_excel_rgb(qt_item.background().color().name()),
            fill_type="solid"
        )
        excel_cell.fill = fill
        
        # Alignment
        align = Alignment(
            horizontal=StyleConverter.qt_to_excel_alignment(qt_item.textAlignment()),
            vertical="center",
            wrap_text=True
        )
        excel_cell.alignment = align
    
    @staticmethod
    def hex_to_excel_rgb(hex_color: str) -> str:
        """Convert hex color to Excel RGB format"""
        hex_color = hex_color.lstrip('#')
        if len(hex_color) == 3:
            hex_color = ''.join([c * 2 for c in hex_color])
        return f"00{hex_color[4:6]}{hex_color[2:4]}{hex_color[0:2]}" # Corrected order for ARGB
    
    @staticmethod
    def qt_to_excel_alignment(qt_alignment: int) -> str:
        """Convert Qt alignment to Excel alignment string"""
        if qt_alignment & Qt.AlignLeft:
            return "left"
        elif qt_alignment & Qt.AlignRight:
            return "right"
        elif qt_alignment & Qt.AlignHCenter:
            return "center"
        elif qt_alignment & Qt.AlignJustify:
            return "justify"
        return "left"
# --- End of StyleConverter ---

# --- Start of ExcelTableWidget ---
class ExcelTableWidget(QTableWidget):
    """Enhanced QTableWidget with Excel-like features and style preservation"""
    
    def __init__(self):
        super().__init__()
        self.setup_ui()
        self.merged_cells = []
    
    def setup_ui(self):
        """Initialize table with Excel-like appearance"""
        self.setEditTriggers(QAbstractItemView.AllEditTriggers)
        self.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.setSelectionMode(QAbstractItemView.ContiguousSelection)
        self.setAlternatingRowColors(False)
        self.setStyleSheet("""
            QTableWidget {
                background-color: white;
                gridline-color: #d0d0d0;
            }
            QTableWidget::item {
                padding: 3px;
            }
            QTableWidget::item:selected {
                background-color: #0078d7;
                color: white;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 5px;
                border: 1px solid #d0d0d0;
            }
        """)
        self.horizontalHeader().setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.verticalHeader().setDefaultSectionSize(24)
        self.verticalHeader().setVisible(False) # Usually row numbers are not shown like this
    
    def load_excel_sheet(self, worksheet: Worksheet):
        """Load Excel worksheet into the table while preserving all styles"""
        try:
            self.clear()
            self.setRowCount(min(worksheet.max_row, MAX_ROWS_PREVIEW))
            self.setColumnCount(min(worksheet.max_column, MAX_COLS_PREVIEW))
            
            headers = []
            for col_idx in range(1, self.columnCount() + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                headers.append(str(cell.value) if cell.value else f"Column {col_idx}")
            self.setHorizontalHeaderLabels(headers)
            
            for row_idx_table in range(self.rowCount()): # 0-indexed for table widget
                for col_idx_table in range(self.columnCount()): # 0-indexed for table widget
                    excel_row = row_idx_table + 1
                    excel_col = col_idx_table + 1

                    cell = worksheet.cell(row=excel_row, column=excel_col)
                    value = self.format_cell_value(cell.value)
                    
                    item = QTableWidgetItem(value)
                    self.apply_excel_style(cell, item)
                    self.setItem(row_idx_table, col_idx_table, item)
            
            self.merged_cells = []
            for merge_range in worksheet.merged_cells.ranges:
                self.setSpan(
                    merge_range.min_row - 1,
                    merge_range.min_col - 1,
                    merge_range.max_row - merge_range.min_row + 1,
                    merge_range.max_col - merge_range.min_col + 1
                )
                self.merged_cells.append((
                    merge_range.min_row - 1,
                    merge_range.min_col - 1,
                    merge_range.max_row - merge_range.min_row + 1,
                    merge_range.max_col - merge_range.min_col + 1
                ))
            
            self.resizeColumnsToContents()
            return True
            
        except Exception as e:
            logger.error(f"Error loading Excel sheet into table: {str(e)}")
            return False
    
    def apply_excel_style(self, excel_cell, qt_item):
        style = StyleConverter.excel_to_qt(excel_cell)
        font = QFont(style.font_name, style.font_size)
        font.setBold(style.bold); font.setItalic(style.italic); font.setUnderline(style.underline)
        qt_item.setFont(font)
        qt_item.setForeground(QColor(style.text_color)); qt_item.setBackground(QColor(style.bg_color))
        alignment = 0
        if style.h_align == "left": alignment |= Qt.AlignLeft
        elif style.h_align == "right": alignment |= Qt.AlignRight
        elif style.h_align == "center": alignment |= Qt.AlignHCenter
        elif style.h_align == "justify": alignment |= Qt.AlignJustify
        if style.v_align == "top": alignment |= Qt.AlignTop
        elif style.v_align == "center": alignment |= Qt.AlignVCenter
        elif style.v_align == "bottom": alignment |= Qt.AlignBottom
        else: alignment |= Qt.AlignVCenter
        qt_item.setTextAlignment(alignment)
    
    def format_cell_value(self, value) -> str:
        if value is None: return ""
        if isinstance(value, datetime): return value.strftime("%d/%m/%Y %H:%M")
        if isinstance(value, (int, float)): return f"{value:,}"
        return str(value)

    def add_row(self):
        rc = self.rowCount(); self.insertRow(rc)
        if self.columnCount() == 0: self.insertColumn(0); self.setHorizontalHeaderItem(0, QTableWidgetItem("Column 1"))
    def add_column(self):
        cc = self.columnCount(); self.insertColumn(cc)
        self.setHorizontalHeaderItem(cc, QTableWidgetItem(f"Column {cc + 1}"))
        if self.rowCount() == 0: self.insertRow(0)
    def delete_selected_rows(self):
        sr = sorted(list(set(i.row() for i in self.selectedIndexes())), reverse=True)
        for r_idx in sr: self.removeRow(r_idx)
    def delete_selected_columns(self):
        sc = sorted(list(set(i.column() for i in self.selectedIndexes())), reverse=True)
        for c_idx in sc: self.removeColumn(c_idx)
# --- End of ExcelTableWidget ---

class PDFGenerator:
    """Handles high-quality PDF export with modern styling"""
    
    def __init__(self, table: ExcelTableWidget, client_data: ClientData, settings: PDFExportSettings,
                 current_sheet_title: str, sheet_images_data: Dict[str, List],
                 sheet_headers_footers_data: Dict[str, Dict[str, Dict[str, Dict[str, Optional[str]]]]]):
        self.table = table
        self.client_data = client_data
        self.settings = settings
        self.current_sheet_title = current_sheet_title
        self.sheet_images_data = sheet_images_data
        self.sheet_hf_data = sheet_headers_footers_data
        self.styles = getSampleStyleSheet()
        
        self.register_fonts()
        self.create_styles()
    
    def register_fonts(self):
        try:
            pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
            pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
            pdfmetrics.registerFont(TTFont('Arial-Italic', 'ariali.ttf'))
            pdfmetrics.registerFont(TTFont('Arial-BoldItalic', 'arialbi.ttf'))
        except: pass
    
    def create_styles(self):
        self.title_style = ParagraphStyle('Title', parent=self.styles['Heading1'], fontName='Helvetica-Bold', fontSize=16, leading=18, spaceAfter=12, alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'))
        self.client_style = ParagraphStyle('Client', parent=self.styles['BodyText'], fontName='Helvetica', fontSize=10, leading=12, spaceAfter=6, textColor=colors.HexColor('#34495e'))
        self.table_header_style = ParagraphStyle('TableHeader', parent=self.styles['BodyText'], fontName='Helvetica-Bold', fontSize=9, leading=10, alignment=TA_CENTER, textColor=colors.white, backColor=colors.HexColor('#3498db'))
        self.table_cell_style = ParagraphStyle('TableCell', parent=self.styles['BodyText'], fontName='Helvetica', fontSize=8, leading=9, textColor=colors.black)
        self.footer_style = ParagraphStyle('Footer', parent=self.styles['BodyText'], fontName='Helvetica', fontSize=8, leading=9, alignment=TA_CENTER, textColor=colors.HexColor('#7f8c8d'))
    
    def generate(self, file_path: str) -> Tuple[bool, str]:
        try:
            doc = SimpleDocTemplate(file_path, pagesize=self.settings.page_size, leftMargin=self.settings.margins[0], rightMargin=self.settings.margins[1], topMargin=self.settings.margins[2], bottomMargin=self.settings.margins[3])
            elements = []
            has_excel_header, has_excel_footer = False, False
            if self.current_sheet_title and self.current_sheet_title in self.sheet_hf_data:
                sheet_hf = self.sheet_hf_data[self.current_sheet_title]
                if any(k in sheet_hf for k in ['odd_header', 'even_header', 'first_header']): has_excel_header = True
                if any(k in sheet_hf for k in ['odd_footer', 'even_footer', 'first_footer']): has_excel_footer = True
            
            if self.settings.header and not has_excel_header: elements.extend(self.create_header())
            elements.append(Paragraph(f"<u>Devis {self.client_data.project_id}</u>" if self.client_data.project_id else "<u>Devis</u>", self.title_style))
            elements.append(Spacer(1, 12))
            elements.extend(self.create_client_info())
            elements.append(Spacer(1, 15))

            if self.current_sheet_title and self.current_sheet_title in self.sheet_images_data:
                images_on_sheet = self.sheet_images_data[self.current_sheet_title]
                if images_on_sheet:
                    heading_style_name = 'Heading3' if 'Heading3' in self.styles else 'h3'
                    elements.append(Paragraph("<u>Sheet Images:</u>", self.styles[heading_style_name]))
                    elements.append(Spacer(1, 0.1*inch))
                    for i, img_obj in enumerate(images_on_sheet):
                        try:
                            img_data = img_obj._data() # Corrected from .data()
                            img_flowable = Image(io.BytesIO(img_data))
                            img_flowable.drawWidth = 2 * inch
                            img_flowable.drawHeight = (img_flowable.imageHeight / img_flowable.imageWidth) * (2 * inch) if img_flowable.imageWidth else 1 * inch
                            img_flowable.preserveAspectRatio = True
                            elements.append(img_flowable)
                            elements.append(Spacer(1, 0.1*inch))
                            logger.info(f"Added image {i+1} (Type: {type(img_obj)}) from sheet '{self.current_sheet_title}' to PDF.")
                        except Exception as e: logger.error(f"Could not process image from sheet '{self.current_sheet_title}': {e}")
                    elements.append(Spacer(1, 0.2*inch))
            
            table_data, col_widths = self.prepare_table_data()
            if table_data:
                main_table = Table(table_data, colWidths=col_widths, repeatRows=1 if self.settings.repeat_headers else 0)
                main_table.setStyle(self.create_table_style(len(table_data), len(table_data[0])))
                elements.append(main_table)
            else:
                logger.warning(f"No table data for sheet '{self.current_sheet_title}'.")
                if not elements: return False, f"No content for sheet '{self.current_sheet_title}'"
            
            if self.settings.footer and not has_excel_footer:
                elements.append(Spacer(1, 10))
                elements.extend(self.create_footer())
            
            doc.build(elements, onFirstPage=self.add_page_decorations, onLaterPages=self.add_page_decorations)
            return True, "PDF generated successfully"
        except Exception as e:
            logger.error(f"PDF generation failed: {str(e)}", exc_info=True)
            return False, str(e)
    
    def create_header(self) -> List[Flowable]:
        header_elements = []
        header_table = Table([[self.create_company_info(),self.create_document_info()]], colWidths=['70%', '30%'])
        header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('BOTTOMPADDING', (0,0), (-1,-1), 10), ('LEFTPADDING', (0,0), (0,0), 0), ('RIGHTPADDING', (1,0), (1,0), 0)]))
        header_elements.append(header_table)
        header_elements.append(Spacer(1, 10))
        return header_elements
    
    def create_company_info(self) -> Paragraph:
        company_info = [f"<b>{self.client_data.company or 'Société'}</b>", self.client_data.address or "Adresse", f"Tél: {self.client_data.phone or 'N/A'}", f"Email: {self.client_data.email or 'N/A'}"]
        return Paragraph("<br/>".join(company_info), self.client_style)
    
    def create_document_info(self) -> Paragraph:
        doc_info = [f"<b>Devis N° {self.client_data.project_id or 'XXXX'}</b>", f"Date: {datetime.now().strftime('%d/%m/%Y')}", f"Client: {self.client_data.name or 'N/A'}"]
        return Paragraph("<br/>".join(doc_info), self.client_style)
    
    def create_client_info(self) -> List[Flowable]:
        client_elements = []
        client_table = Table([["Client:", self.client_data.name or "N/A"],["Société:", self.client_data.company or "N/A"],["Projet:", self.client_data.project or "N/A"],["Date:", datetime.now().strftime("%d/%m/%Y")],["Prix Total:", f"{self.client_data.price:,.2f} {self.client_data.currency}"]], colWidths=[3*cm, 10*cm])
        client_table.setStyle(TableStyle([('BACKGROUND', (0,0), (0,-1), colors.HexColor("#f8f9fa")),('TEXTCOLOR', (0,0), (-1,-1), colors.black),('ALIGN', (0,0), (0,-1), 'RIGHT'),('ALIGN', (1,0), (1,-1), 'LEFT'),('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'), ('FONTSIZE', (0,0),(-1,-1), 9),('GRID', (0,0),(-1,-1), 0.5, colors.HexColor("#e0e0e0")),('VALIGN', (0,0),(-1,-1), 'MIDDLE'),('BOTTOMPADDING', (0,0),(-1,-1), 6), ('LEFTPADDING', (0,0),(-1,-1), 4), ('RIGHTPADDING', (0,0),(-1,-1), 4)]))
        client_elements.append(client_table)
        return client_elements
    
    def prepare_table_data(self) -> Tuple[List[List[Union[str, Paragraph]]], List[float]]:
        if not self.table: return [],[]
        rows = min(self.table.rowCount(), MAX_ROWS_PREVIEW)
        cols = min(self.table.columnCount(), MAX_COLS_PREVIEW)
        if rows == 0 or cols == 0: return [], []

        data = []
        col_widths = []
        headers = []
        for col in range(cols):
            header = self.table.horizontalHeaderItem(col)
            header_text = header.text() if header else f"Col {col+1}"
            headers.append(Paragraph(header_text, self.table_header_style))
            col_widths.append(self.table.columnWidth(col) / 10.0 if self.table.columnWidth(col) > 0 else 1.5*inch)
        data.append(headers)
        
        for row in range(rows):
            row_data = []
            for col in range(cols):
                item = self.table.item(row, col)
                if item:
                    cell_style = self.table_cell_style.clone('TableCell_Specific')
                    text_color = item.foreground().color().name()
                    if text_color != "#000000": cell_style.textColor = colors.HexColor(text_color)
                    align = item.textAlignment()
                    if align & Qt.AlignLeft: cell_style.alignment = TA_LEFT
                    elif align & Qt.AlignRight: cell_style.alignment = TA_RIGHT
                    elif align & Qt.AlignHCenter: cell_style.alignment = TA_CENTER
                    elif align & Qt.AlignJustify: cell_style.alignment = TA_JUSTIFY
                    row_data.append(Paragraph(item.text(), cell_style))
                else: row_data.append("")
            data.append(row_data)
        return data, col_widths
    
    def create_table_style(self, num_rows: int, num_cols: int) -> TableStyle:
        if num_rows == 0 or num_cols == 0: return TableStyle([])
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#3498db")), ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('ALIGN', (0,0), (-1,0), 'CENTER'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9), ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e0e0e0")),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#f8f9fa")]),
            ('LEFTPADDING', (0,0), (-1,-1), 4), ('RIGHTPADDING', (0,0), (-1,-1), 4), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
        ])
        if self.table:
            for r in range(1, num_rows):
                 for c in range(num_cols):
                    if r >= self.table.rowCount() or c >= self.table.columnCount(): continue
                    item = self.table.item(r-1, c)
                    if item:
                        bg_color_name = item.background().color().name()
                        if bg_color_name != "#ffffff":
                            style.add('BACKGROUND', (c,r), (c,r), colors.HexColor(bg_color_name))
                        row_span = self.table.rowSpan(r-1, c)
                        col_span = self.table.columnSpan(r-1, c)
                        if row_span > 1 or col_span > 1:
                            style.add('SPAN', (c,r), (c+col_span-1, r+row_span-1))
        return style
    
    def create_footer(self) -> List[Flowable]:
        footer_elements = []
        if self.client_data.notes:
            notes = Paragraph(f"<b>Notes:</b><br/>{self.client_data.notes}", ParagraphStyle('Notes',parent=self.styles['BodyText'],fontName='Helvetica',fontSize=8,leading=9,textColor=colors.HexColor("#7f8c8d")))
            footer_elements.append(notes)
            footer_elements.append(Spacer(1, 10))
        footer_text = Paragraph(f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} | {self.client_data.company or 'Société'}", self.footer_style)
        footer_elements.append(footer_text)
        return footer_elements

    def _draw_hf_section(self, canvas: canvas.Canvas, doc, section_content: Dict, is_header: bool, default_hf_height: float, font_name: str, font_size: int):
        canvas.setFont(font_name, font_size)
        y_position = 0
        font_ascent_approx = font_size * 0.8
        if is_header:
            y_position = doc.height + doc.topMargin - default_hf_height + font_ascent_approx * 0.5
        else:
            y_position = doc.bottomMargin + font_ascent_approx * 0.5

        positions = {
            "left": doc.leftMargin + 0.1*inch,
            "center": doc.width / 2 + doc.leftMargin,
            "right": doc.width + doc.leftMargin - 0.1*inch
        }

        for part_key in ["left", "center", "right"]:
            part_data = section_content.get(part_key)
            if not part_data: continue
            text_content = part_data.get("text")
            if text_content:
                text_to_draw = text_content
                if "&G" in text_to_draw:
                    logger.info(f"Header/Footer image reference mark (&G) found in {part_key} of {'header' if is_header else 'footer'}: '{text_to_draw}'. Actual display requires rID/&G parsing and image data resolution.")

                page_number_str = str(canvas.getPageNumber())
                total_pages_str = str(getattr(doc, '_pageNumber', canvas.getPageNumber()))
                text_to_draw = text_to_draw.replace("&[Page]", page_number_str).replace("&P", page_number_str)
                text_to_draw = text_to_draw.replace("&[Pages]", total_pages_str).replace("&N", total_pages_str)
                text_to_draw = text_to_draw.replace("&[Date]", datetime.now().strftime("%Y-%m-%d")).replace("&D", datetime.now().strftime("%Y-%m-%d"))
                text_to_draw = text_to_draw.replace("&[Time]", datetime.now().strftime("%H:%M:%S")).replace("&T", datetime.now().strftime("%H:%M:%S"))
                text_to_draw = text_to_draw.replace("&[File]", Path(doc.filename).name if doc.filename else "Document").replace("&F", Path(doc.filename).name if doc.filename else "Document")
                text_to_draw = text_to_draw.replace("&[Tab]", self.current_sheet_title or "Sheet").replace("&A", self.current_sheet_title or "Sheet")
                text_to_draw = text_to_draw.replace("&B", "").replace("&I", "")
                text_to_draw = text_to_draw.replace("&G", "").strip()

                canvas.setFillColor(colors.black)
                if part_key == "left": canvas.drawString(positions[part_key], y_position, text_to_draw)
                elif part_key == "center": canvas.drawCentredString(positions[part_key], y_position, text_to_draw)
                elif part_key == "right": canvas.drawRightString(positions[part_key], y_position, text_to_draw)
    
    def add_page_decorations(self, canvas: canvas.Canvas, doc):
        canvas.saveState()
        if self.settings.watermark:
            canvas.setFont('Helvetica', 60) # Changed from Arial
            canvas.setFillColor(colors.HexColor("#f0f0f0"))
            canvas.rotate(45)
            page_width, page_height = doc.pagesize
            canvas.drawString(page_width * 0.25, page_height * 0.25, self.settings.watermark)
            canvas.rotate(-45)

        default_page_num_text = f"Page {canvas.getPageNumber()}"
        excel_hf_applied_for_page = False
        hf_font_name = "Helvetica"
        hf_font_size = 8

        if self.current_sheet_title and self.current_sheet_title in self.sheet_hf_data:
            sheet_hf_content = self.sheet_hf_data[self.current_sheet_title]
            page_num = canvas.getPageNumber()
            header_type_to_use, footer_type_to_use = None, None

            if page_num == 1 and sheet_hf_content.get("first_header"): header_type_to_use = "first_header"
            elif page_num % 2 == 0 and sheet_hf_content.get("even_header"): header_type_to_use = "even_header"
            elif sheet_hf_content.get("odd_header"): header_type_to_use = "odd_header"

            if page_num == 1 and sheet_hf_content.get("first_footer"): footer_type_to_use = "first_footer"
            elif page_num % 2 == 0 and sheet_hf_content.get("even_footer"): footer_type_to_use = "even_footer"
            elif sheet_hf_content.get("odd_footer"): footer_type_to_use = "odd_footer"

            if header_type_to_use and header_type_to_use in sheet_hf_content:
                header_data = sheet_hf_content[header_type_to_use]
                self._draw_hf_section(canvas, doc, header_data, True, PDF_HEADER_HEIGHT, hf_font_name, hf_font_size)
                if any(header_data.get(part, {}).get('text') for part in ["left", "center", "right"]):
                    excel_hf_applied_for_page = True

            if footer_type_to_use and footer_type_to_use in sheet_hf_content:
                footer_data = sheet_hf_content[footer_type_to_use]
                self._draw_hf_section(canvas, doc, footer_data, False, PDF_FOOTER_HEIGHT, hf_font_name, hf_font_size)
                if any(footer_data.get(part, {}).get('text') for part in ["left", "center", "right"]):
                    excel_hf_applied_for_page = True

        if self.settings.footer and not excel_hf_applied_for_page:
            canvas.setFont(hf_font_name, hf_font_size)
            canvas.setFillColor(colors.HexColor("#7f8c8d"))
            page_num_y_pos = doc.bottomMargin + (PDF_FOOTER_HEIGHT - hf_font_size * 0.8) / 2
            canvas.drawRightString(doc.width + doc.leftMargin - 0.1*inch, page_num_y_pos, default_page_num_text)
        canvas.restoreState()

# --- Start of ExcelEditor ---
class ExcelEditor(QDialog):
    """Main Excel editor application with modern UI"""
    
    def __init__(self, file_path: str = "", parent=None):
        super().__init__(parent)
        self.file_path: Optional[str] = file_path if file_path else None # Store initial file path
        self.excel_model = ExcelTableModel() # Instantiate without file_path
        self.pdf_settings = PDFExportSettings()
        
        self.setup_ui() # Creates self.table, self.sheet_combo, self.file_label etc.
        self.setup_connections()
        
        if self.file_path: # If a file_path was provided to constructor
            self.load_file(self.file_path)
    
    def setup_ui(self):
        """Initialize the main UI"""
        self.setWindowTitle("Ultimate Excel Editor")
        self.setMinimumSize(1280, 800)
        self.resize(1366, 768)
        
        # Apply modern style
        self.setStyle(QStyleFactory.create("Fusion"))
        
        # Dark palette for modern look
        palette = self.palette()
        palette.setColor(QPalette.Window, QColor(53, 53, 53))
        palette.setColor(QPalette.WindowText, Qt.white)
        palette.setColor(QPalette.Base, QColor(25, 25, 25))
        palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
        palette.setColor(QPalette.ToolTipBase, Qt.white)
        palette.setColor(QPalette.ToolTipText, Qt.white)
        palette.setColor(QPalette.Text, Qt.white)
        palette.setColor(QPalette.Button, QColor(53, 53, 53))
        palette.setColor(QPalette.ButtonText, Qt.white)
        palette.setColor(QPalette.BrightText, Qt.red)
        palette.setColor(QPalette.Link, QColor(42, 130, 218))
        palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        palette.setColor(QPalette.HighlightedText, Qt.black)
        self.setPalette(palette)
        
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(5, 5, 5, 5)
        main_layout.setSpacing(5)

        # Initialize components that are dependencies first
        self.table_panel = self.create_table_panel() # Initializes self.table
        self.client_panel = self.create_client_panel()

        # Create toolbar (depends on self.table)
        self.create_toolbar(main_layout) # Adds toolbar to main_layout

        # Main content area
        content_splitter = QSplitter(Qt.Horizontal)
        content_splitter.addWidget(self.client_panel)
        content_splitter.addWidget(self.table_panel) # self.table_panel now already created
        content_splitter.setStretchFactor(0, 1) # client_panel
        content_splitter.setStretchFactor(1, 3) # table_panel
        main_layout.addWidget(content_splitter)

        # Create status bar (usually at the bottom)
        self.create_status_bar(main_layout) # Adds status_bar to main_layout

        # Progress bar (below status bar)
        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(20)
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)
    
    def create_toolbar(self, parent_layout):
        """Create main application toolbar"""
        toolbar = QToolBar()
        toolbar.setIconSize(QSize(24, 24))
        toolbar.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)
        toolbar.setMovable(False)
        
        # File actions
        open_action = QAction(QIcon.fromTheme("document-open"), "Open", self)
        open_action.setShortcut(QKeySequence.Open)
        open_action.triggered.connect(self.open_file_dialog)
        toolbar.addAction(open_action)
        
        save_action = QAction(QIcon.fromTheme("document-save"), "Save", self)
        save_action.setShortcut(QKeySequence.Save)
        save_action.triggered.connect(self.save_file)
        toolbar.addAction(save_action)
        
        toolbar.addSeparator()
        
        # PDF export
        pdf_action = QAction(QIcon.fromTheme("document-export"), "Export PDF", self)
        pdf_action.setShortcut("Ctrl+E")
        pdf_action.triggered.connect(self.export_pdf)
        toolbar.addAction(pdf_action)
        
        # Settings
        settings_action = QAction(QIcon.fromTheme("configure"), "Settings", self)
        settings_action.triggered.connect(self.show_settings)
        toolbar.addAction(settings_action)
        
        toolbar.addSeparator()
        
        # Table actions
        add_row_action = QAction(QIcon.fromTheme("list-add"), "Add Row", self)
        add_row_action.triggered.connect(self.table.add_row)
        toolbar.addAction(add_row_action)
        
        add_col_action = QAction(QIcon.fromTheme("list-add"), "Add Column", self)
        add_col_action.triggered.connect(self.table.add_column)
        toolbar.addAction(add_col_action)
        
        del_row_action = QAction(QIcon.fromTheme("list-remove"), "Delete Row", self)
        del_row_action.triggered.connect(self.table.delete_selected_rows)
        toolbar.addAction(del_row_action)
        
        del_col_action = QAction(QIcon.fromTheme("list-remove"), "Delete Column", self)
        del_col_action.triggered.connect(self.table.delete_selected_columns)
        toolbar.addAction(del_col_action)
        
        parent_layout.addWidget(toolbar)
    
    def create_status_bar(self, parent_layout):
        """Create status bar at bottom of window"""
        status_bar = QFrame()
        status_bar.setFrameShape(QFrame.StyledPanel)
        status_bar.setStyleSheet("background-color: #3a3a3a; border-top: 1px solid #505050;")
        status_bar.setFixedHeight(30)
        
        layout = QHBoxLayout(status_bar)
        layout.setContentsMargins(10, 0, 10, 0)
        
        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("color: #aaaaaa;")
        layout.addWidget(self.status_label)
        
        layout.addStretch()
        
        self.file_label = QLabel("No file loaded")
        self.file_label.setStyleSheet("color: #7f8c8d; font-style: italic;")
        layout.addWidget(self.file_label)
        
        parent_layout.addWidget(status_bar)
    
    def create_client_panel(self) -> QWidget:
        """Create panel for client information"""
        panel = QGroupBox("Client Information")
        panel.setMaximumWidth(350)
        
        layout = QFormLayout(panel)
        layout.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        
        # Client name
        self.client_name = QLineEdit()
        self.client_name.setPlaceholderText("Client name")
        layout.addRow("Name:", self.client_name)
        
        # Company
        self.client_company = QLineEdit()
        self.client_company.setPlaceholderText("Company name")
        layout.addRow("Company:", self.client_company)
        
        # Address
        self.client_address = QTextEdit()
        self.client_address.setMaximumHeight(60)
        self.client_address.setPlaceholderText("Address")
        layout.addRow("Address:", self.client_address)
        
        # Contact info
        contact_layout = QHBoxLayout()
        self.client_phone = QLineEdit()
        self.client_phone.setPlaceholderText("Phone")
        contact_layout.addWidget(self.client_phone)
        
        self.client_email = QLineEdit()
        self.client_email.setPlaceholderText("Email")
        contact_layout.addWidget(self.client_email)
        
        layout.addRow("Contact:", contact_layout)
        
        # Project info
        self.project_name = QLineEdit()
        self.project_name.setPlaceholderText("Project name")
        layout.addRow("Project:", self.project_name)
        
        self.project_id = QLineEdit()
        self.project_id.setPlaceholderText("Project ID")
        layout.addRow("Project ID:", self.project_id)
        
        # Price
        price_layout = QHBoxLayout()
        self.project_price = QDoubleSpinBox()
        self.project_price.setRange(0, 99999999)
        self.project_price.setPrefix("€ ")
        price_layout.addWidget(self.project_price)
        
        self.currency_combo = QComboBox()
        self.currency_combo.addItems(["€", "$", "£", "¥", "₺"])
        price_layout.addWidget(self.currency_combo)
        
        layout.addRow("Price:", price_layout)
        
        # Notes
        self.notes = QTextEdit()
        self.notes.setMaximumHeight(80)
        self.notes.setPlaceholderText("Additional notes")
        layout.addRow("Notes:", self.notes)
        
        # Logo
        self.logo_button = QPushButton("Select Logo...")
        self.logo_button.clicked.connect(self.select_logo)
        layout.addRow("Logo:", self.logo_button)
        
        return panel
    
    def create_table_panel(self) -> QWidget:
        """Create panel for Excel table editing"""
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Sheet selector
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(200)
        layout.addWidget(self.sheet_combo)
        
        # Table widget
        self.table = ExcelTableWidget()
        layout.addWidget(self.table)
        
        return panel
    
    def setup_connections(self):
        """Setup signal connections between UI elements"""
        # Client data changes
        self.client_name.textChanged.connect(self.update_client_data)
        self.client_company.textChanged.connect(self.update_client_data)
        self.client_address.textChanged.connect(self.update_client_data)
        self.client_phone.textChanged.connect(self.update_client_data)
        self.client_email.textChanged.connect(self.update_client_data)
        self.project_name.textChanged.connect(self.update_client_data)
        self.project_id.textChanged.connect(self.update_client_data)
        self.project_price.valueChanged.connect(self.update_client_data)
        self.currency_combo.currentTextChanged.connect(self.update_client_data)
        self.notes.textChanged.connect(self.update_client_data)
        
        # Sheet selection
        self.sheet_combo.currentTextChanged.connect(self.change_sheet)
        
        # Table changes
        self.table.itemChanged.connect(self.mark_as_modified)
    
    def load_file(self, file_path: str):
        """Load Excel file into editor, updating model and UI."""
        self.set_progress(True, f"Loading file: {QFileInfo(file_path).fileName()}...")

        if self.excel_model.load_workbook(file_path):
            self.file_path = file_path
            self.file_label.setText(QFileInfo(file_path).fileName())
            self.sheet_combo.clear()
            self.sheet_combo.addItems(self.excel_model.sheets)
            self.load_client_data()
            if self.excel_model.current_sheet:
                self.table.load_excel_sheet(self.excel_model.current_sheet)
            else:
                self.table.clearContents()
                self.table.setRowCount(0)
                self.table.setColumnCount(0)
            self.set_progress(False)
            self.update_status("File loaded successfully")
            logger.info(f"Successfully loaded and UI updated for: {file_path}")
            return True
        else:
            error_msg = self.excel_model.load_error_message or f"Unknown error loading file: {file_path}"
            QMessageBox.critical(self, "Error Loading File", error_msg)
            self.file_path = None
            self.file_label.setText("No file loaded")
            self.sheet_combo.clear()
            self.table.clearContents()
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.client_name.clear(); self.client_company.clear(); self.client_address.clear()
            self.client_phone.clear(); self.client_email.clear(); self.project_name.clear()
            self.project_id.clear(); self.project_price.setValue(0); self.notes.clear()
            self.set_progress(False)
            self.update_status("Failed to load file", "red")
            logger.error(f"Failed to load and update UI for: {file_path}")
            return False
    
    def load_client_data(self):
        client = self.excel_model.client_data
        self.client_name.setText(client.name)
        self.client_company.setText(client.company)
        self.client_address.setPlainText(client.address)
        self.client_phone.setText(client.phone)
        self.client_email.setText(client.email)
        self.project_name.setText(client.project)
        self.project_id.setText(client.project_id)
        self.project_price.setValue(client.price)
        self.currency_combo.setCurrentText(client.currency)
        self.notes.setPlainText(client.notes)
    
    def update_client_data(self):
        self.excel_model.client_data = ClientData(
            name=self.client_name.text(), company=self.client_company.text(),
            address=self.client_address.toPlainText(), phone=self.client_phone.text(),
            email=self.client_email.text(), project=self.project_name.text(),
            project_id=self.project_id.text(), price=self.project_price.value(),
            currency=self.currency_combo.currentText(), notes=self.notes.toPlainText(),
            logo_path=self.excel_model.client_data.logo_path
        )
        self.mark_as_modified()
    
    def mark_as_modified(self):
        self.excel_model.is_modified = True
        self.update_status("Modified", "orange")
    
    def update_status(self, message: str, color: str = "green"):
        self.status_label.setText(message)
        self.status_label.setStyleSheet(f"color: {color};")
    
    def set_progress(self, visible: bool, message: str = ""):
        self.progress_bar.setVisible(visible)
        if visible:
            self.progress_bar.setRange(0, 0) # Indeterminate
            self.update_status(message, "blue")
        else:
            self.progress_bar.setRange(0, 1) # Stop/hide
    
    def open_file_dialog(self):
        if self.excel_model.is_modified:
            reply = QMessageBox.question(self, "Unsaved Changes", "You have unsaved changes. Save before opening new file?", QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel, QMessageBox.Save)
            if reply == QMessageBox.Save:
                if not self.save_file(): return
            elif reply == QMessageBox.Cancel: return
        
        file_path, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls *.xlsm);;All Files (*)")
        if file_path: self.load_file(file_path)
    
    def save_file(self) -> bool:
        if not self.file_path:
            return self.save_file_as()
        if not self.excel_model.workbook or not self.excel_model.current_sheet:
            QMessageBox.warning(self, "Error", "No workbook or sheet loaded to save.")
            return False

        self.set_progress(True, "Saving file...")
        self.update_client_data()  # Ensure client data in model is up-to-date

        # Write QTableWidget data back to the openpyxl worksheet
        worksheet = self.excel_model.current_sheet
        for r in range(self.table.rowCount()):
            for c in range(self.table.columnCount()):
                item = self.table.item(r, c)
                if item:
                    cell_value_str = item.text()
                    # Attempt to convert back to numeric if possible, otherwise save as string
                    # More sophisticated type handling might be needed for dates, specific number formats etc.
                    try:
                        # Try float conversion for cells that might contain numbers
                        # Excel itself stores numbers as floats mostly.
                        # Remove commas for thousands separator if used in display
                        cleaned_value_str = cell_value_str.replace(',', '')
                        cell_value = float(cleaned_value_str)
                    except ValueError:
                        # If not a float, try int
                        try:
                            cell_value = int(cleaned_value_str)
                        except ValueError:
                            # Fallback to string if not float or int
                            cell_value = cell_value_str

                    # Apply value to openpyxl cell
                    # worksheet.cell(row=r + 1, column=c + 1).value = item.text() # Original simple save
                    excel_cell = worksheet.cell(row=r + 1, column=c + 1)
                    excel_cell.value = cell_value

                    # Optionally, re-apply styles from QTableWidgetItem back to excel_cell
                    # This is complex if styles can be modified in QTableWidget.
                    # For now, assume styles are preserved from load or managed by openpyxl.
                    # StyleConverter.qt_to_excel(item, excel_cell) # If style changes in table are to be saved
                else:
                    # If QTableWidgetItem is None (e.g. if cells were cleared in UI)
                    worksheet.cell(row=r + 1, column=c + 1).value = None


        try:
            self.excel_model.workbook.save(self.file_path)
            self.excel_model.is_modified = False
            self.set_progress(False)
            self.update_status("File saved successfully")
            return True
        except ReadOnlyWorkbookException:
            logger.error(f"Cannot save read-only workbook: {self.file_path}")
            QMessageBox.critical(self, "Error", f"Cannot save read-only workbook: {self.file_path}\nPlease use 'Save As...' to save a copy.")
            self.set_progress(False)
            return False
        except Exception as e:
            logger.error(f"Error saving file: {str(e)}")
            QMessageBox.critical(self, "Error", f"Error saving file:\n{str(e)}")
            self.set_progress(False)
            return False
    
    def save_file_as(self) -> bool:
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)")
        if file_path:
            self.file_path = file_path
            self.file_label.setText(QFileInfo(file_path).fileName())
            return self.save_file()
        return False
    
    def change_sheet(self, sheet_name: str):
        if not sheet_name or not self.excel_model.workbook: return
        if self.excel_model.current_sheet and sheet_name == self.excel_model.current_sheet.title: return
            
        if self.excel_model.is_modified: # Check for unsaved changes
            reply = QMessageBox.question(self, "Unsaved Changes", "You have unsaved changes. Save before switching sheets?", QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel, QMessageBox.Save)
            if reply == QMessageBox.Save:
                if not self.save_file(): return
            elif reply == QMessageBox.Cancel:
                self.sheet_combo.setCurrentText(self.excel_model.current_sheet.title) # Revert selection
                return
        
        try:
            self.set_progress(True, f"Loading sheet: {sheet_name}")
            self.excel_model.current_sheet = self.excel_model.workbook[sheet_name]
            self.table.load_excel_sheet(self.excel_model.current_sheet)
            self.set_progress(False)
            self.update_status(f"Sheet '{sheet_name}' loaded")
        except Exception as e:
            logger.error(f"Error changing sheet: {str(e)}")
            QMessageBox.critical(self, "Error", f"Error loading sheet:\n{str(e)}")
            self.set_progress(False)
    
    def export_pdf(self):
        if not self.file_path:
            QMessageBox.warning(self, "Warning", "No file loaded to export")
            return
        self.update_client_data()
        default_name = f"{Path(self.file_path).stem}.pdf"
        save_path, _ = QFileDialog.getSaveFileName(self, "Export PDF", default_name, "PDF Files (*.pdf);;All Files (*)")
        if not save_path: return
            
        try:
            self.set_progress(True, "Exporting to PDF...")
            current_sheet_title = self.excel_model.current_sheet.title if self.excel_model.current_sheet else ""
            pdf_gen = PDFGenerator(
                self.table, self.excel_model.client_data, self.pdf_settings,
                current_sheet_title, self.excel_model.sheet_images,
                self.excel_model.sheet_headers_footers
            )
            success, message = pdf_gen.generate(save_path)
            self.set_progress(False)
            if success:
                self.update_status("PDF exported successfully")
                if QMessageBox.question(self, "Export Complete", f"PDF created successfully:\n{save_path}\n\nOpen now?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                    self.open_file(save_path)
            else:
                QMessageBox.critical(self, "Export Error", f"Error exporting PDF:\n{message}")
        except Exception as e:
            logger.error(f"Error exporting PDF: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "Error", f"Error exporting PDF:\n{str(e)}")
            self.set_progress(False)
    
    def select_logo(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Logo Image", "", "Image Files (*.png *.jpg *.jpeg *.bmp);;All Files (*)")
        if file_path:
            self.excel_model.client_data.logo_path = file_path
            self.mark_as_modified()
    
    def show_settings(self):
        QMessageBox.information(self, "Settings", "PDF export settings will be available in the next version")
    
    def open_file(self, file_path: str):
        try:
            if sys.platform == "win32": os.startfile(file_path)
            elif sys.platform == "darwin": os.system(f'open "{file_path}"')
            else: os.system(f'xdg-open "{file_path}"')
        except Exception as e:
            logger.error(f"Could not open file: {str(e)}")
            QMessageBox.warning(self, "Warning", f"Could not open file:\n{str(e)}")

def main():
    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create("Fusion"))
    app.setApplicationName("Ultimate Excel Editor")
    app.setApplicationVersion("1.0")
    app.setWindowIcon(QIcon.fromTheme("x-office-spreadsheet")) # Needs a valid theme or direct path
    
    editor = ExcelEditor()
    # Example: Load a file if passed as a command-line argument
    if len(sys.argv) > 1:
        editor.load_file(sys.argv[1])
    editor.show()
    
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
# --- End of ExcelEditor ---
